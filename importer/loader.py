import openpyxl
import xlrd
import re
from collections import namedtuple
from dateutil import parser

RawRead = namedtuple('RawRead', 'forward_read, reverse_read, sample_name, sample_accession, taxon_id, library_name')


class Spreadsheet:

    @staticmethod
    def new_instance(name, reads=[], accession='', contact='', organisation='', size=0, supplier='', technology='',
                     limit=''):
        result = Spreadsheet()
        result.name = name
        result.reads = reads
        result.accession = accession
        result.contact = contact
        result.organisation = organisation
        result.size = size
        result.supplier = supplier
        result.technology = technology
        result.limit = limit
        return result



class SpreadsheetLoader:

    def __init__(self, file):
        self._file = file
        if file.lower().endswith(".xlsx"):
            self._format = 'xlsx'
            self._workbook = openpyxl.load_workbook(self._file)
            self._sheet = self._workbook.worksheets[0]
        else:
            self._format = 'xls'
            self._workbook = xlrd.open_workbook(self._file)
            self._sheet = self._workbook.sheet_by_index(0)

    def load(self):
        if self._format == 'xlsx':
            return self.load_xlsx()
        else:
            return self.load_xls()

    def load_xls(self):
        result = Spreadsheet()
        data_row = 0
        header_row = 0
        for i in range(self._sheet.nrows):
            if self._sheet.cell_value(i, 0) == 'Study Name':
                result.name = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Supplier Name':
                result.supplier = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Supplier Organisation':
                result.organisation = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Sanger Contact Name':
                result.contact = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Sequencing Technology':
                result.technology = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Study Accession number':
                result.accession = self.__extract_text_value_xls(i, 1)
            if self._sheet.cell_value(i, 0) == 'Total size of files in GBytes':
                result.size = self._sheet.cell_value(i, 1)
            if self._sheet.cell_value(i, 0) == 'Data to be kept until':
                year, month, day, hour, minute, second = xlrd.xldate_as_tuple(self._sheet.cell_value(i, 1),
                                                                              self._workbook.datemode)
                result.limit = "%02d/%02d/%04d" % (day, month, year)
            if self._sheet.cell_value(i, 0) == 'Filename' or self._sheet.cell_value(i, 0) == 'Run Accession':
                data_row = i + 1
                header_row = i
                break
        filename_column = None
        run_accession_column = None
        for i in range(self._sheet.ncols):
            if self._sheet.cell_value(header_row, i) == 'Filename':
                filename_column = i
            if self._sheet.cell_value(header_row, i) == 'Run Accession':
                run_accession_column = i
            if filename_column is not None:
                if self._sheet.cell_value(header_row, i) == 'Mate File':
                    mate_filename_column = i
            if run_accession_column is not None:
                if self._sheet.cell_value(header_row, i) == 'Double-ended Reads':
                    double_ended_reads_column = i
            if self._sheet.cell_value(header_row, i) == 'Sample Name':
                sample_name_column = i
            if self._sheet.cell_value(header_row, i) == 'Sample Accession number':
                sample_accession_column = i
            if self._sheet.cell_value(header_row, i) == 'Taxon ID':
                taxon_id_column = i
            if self._sheet.cell_value(header_row, i) == 'Library Name':
                library_name_column = i
        reads = []
        for i in range(data_row, self._sheet.nrows):
            sample_name = self.__extract_float_value_xls(i, sample_name_column)
            library_name = self.__extract_float_value_xls(i, library_name_column)
            if library_name is None:
                library_name = sample_name
            if filename_column is not None:
                reads.append(RawRead(
                    self.__extract_text_value_xls(i, filename_column),
                    self.__extract_text_value_xls(i, mate_filename_column),
                    sample_name,
                    self.__extract_text_value_xls(i, sample_accession_column),
                    self.__extract_float_value_xls(i, taxon_id_column),
                    library_name))
            if run_accession_column is not None:
                reads.append(RawRead(
                    (self.__extract_text_value_xls(i, run_accession_column)),
                    self.__extract_text_value_xls(i, double_ended_reads_column),
                    sample_name,
                    self.__extract_text_value_xls(i, sample_accession_column),
                    self.__extract_float_value_xls(i, taxon_id_column),
                    library_name))
            reads = self.trim_blank_ends(reads)
        result.reads = reads
        return result

    @staticmethod
    def trim_blank_ends(reads):
        blank_reads = True
        while blank_reads:
            if reads[-1].forward_read == None and reads[-1].reverse_read == None and reads[-1].sample_name == None \
                    and reads[-1].sample_accession == None:
                reads = reads[:-1]
            else:
                blank_reads = False
        return reads

    def load_xlsx(self):
        result = Spreadsheet()
        data_row = 0
        header_row = 0
        for i in range(10):
            if self._sheet.cell(row=i + 1, column=1).value == 'Study Name':
                result.name = self._sheet.cell(row=i + 1, column=2).value
            if self._sheet.cell(row=i + 1, column=1).value == 'Supplier Name':
                result.supplier = self._sheet.cell(row=i + 1, column=2).value
            if self._sheet.cell(row=i + 1, column=1).value == 'Supplier Organisation':
                result.organisation = self._sheet.cell(row=i + 1, column=2).value
            if self._sheet.cell(row=i + 1, column=1).value == 'Sanger Contact Name':
                result.contact = self._sheet.cell(row=i + 1, column=2).value
            if self._sheet.cell(row=i + 1, column=1).value == 'Sequencing Technology':
                result.technology = self._sheet.cell(row=i + 1, column=2).value
            if self._sheet.cell(row=i + 1, column=1).value == 'Study Accession number':
                result.accession = self.__extract_text_value_xlsx(i + 1, 2)
            if self._sheet.cell(row=i + 1, column=1).value == 'Total size of files in GBytes':
                size = self._sheet.cell(row=i + 1, column=2).value
                if isinstance(size,str):
                    size = re.sub(r'[^0-9\.\-]','',size) # Remove non-numeric characters
                result.size = float(size)
            if self._sheet.cell(row=i + 1, column=1).value == 'Data to be kept until':
                limit = self._sheet.cell(row=i + 1, column=2).value
                if isinstance(limit,str):
                    limit = parser.parse(limit)
                result.limit = limit.strftime('%d/%m/%Y')
            if self._sheet.cell(row=i + 1, column=1).value == 'Filename' or self._sheet.cell(row=i + 1, column=1).value == 'Run Accession':
                data_row = i + 2
                header_row = i + 1
                break
        filename_column = None
        run_accession_column = None
        for i in range(self._sheet.max_column):
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Filename':
                filename_column = i +1
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Run Accession':
                run_accession_column = i +1
            if filename_column is not None:
                if self._sheet.cell(row=header_row, column=i + 1).value == 'Mate File':
                    mate_filename_column = i + 1
            if run_accession_column is not None:
                if self._sheet.cell(row=header_row, column=i + 1).value == 'Double-ended Reads':
                    double_ended_reads_column = i + 1
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Sample Name':
                sample_name_column = i + 1
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Sample Accession number':
                sample_accession_column = i + 1
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Taxon ID':
                taxon_id_column = i + 1
            if self._sheet.cell(row=header_row, column=i + 1).value == 'Library Name':
                library_name_column = i + 1
        reads = []
        for i in range(data_row, self._sheet.max_row+1):
            sample_name = self.__extract_float_value_xlsx(i, sample_name_column)
            library_name = self.__extract_float_value_xlsx(i, library_name_column)
            if library_name is None:
                library_name = sample_name
            if filename_column is not None:
                reads.append(RawRead(
                    self.__extract_text_value_xlsx(i, filename_column),
                    self.__extract_text_value_xlsx(i, mate_filename_column),
                    sample_name,
                    self.__extract_text_value_xlsx(i, sample_accession_column),
                    self.__extract_float_value_xlsx(i, taxon_id_column),
                    library_name))
            if run_accession_column is not None:
                reads.append(RawRead(
                    (self.__extract_text_value_xlsx(i, run_accession_column)),
                    self.__extract_text_value_xlsx(i, double_ended_reads_column),
                    sample_name,
                    self.__extract_text_value_xlsx(i, sample_accession_column),
                    self.__extract_float_value_xlsx(i, taxon_id_column),
                    library_name))
            reads = self.trim_blank_ends(reads)
        result.reads = reads
        return result

    def __extract_text_value_xlsx(self, row, column):
        new_data = self._sheet.cell(row=row, column=column).value
        if new_data is None:
            return None
        new_data = str(new_data).strip()
        return None if new_data == '' else new_data

    def __extract_float_value_xlsx(self, row, column):
        if isinstance(self._sheet.cell(row=row, column=column).value, str):
            return self.__extract_text_value_xlsx(row, column)
        value = self._sheet.cell(row=row, column=column).value
        return None if value is None else str(int(value))

    def __extract_text_value_xls(self, row, column):
        new_data = self._sheet.cell_value(row, column)
        if new_data is None:
            return None
        new_data = str(new_data).strip()
        return None if new_data == '' else new_data

    def __extract_float_value_xls(self, row, column):
        if self._sheet.cell_type(row, column) != xlrd.XL_CELL_NUMBER:
            return self.__extract_text_value_xls(row, column)
        return str(int(self._sheet.cell_value(row, column)))

if __name__ == '__main__':
    loader = SpreadsheetLoader('~/Downloads/CHRF_ena_download_template.xls')
    loader.load_xls()