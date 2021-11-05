from os import mkdir, path
from shutil import copyfile

import sys
import openpyxl
import xlwt
import pandas as pd

from importer.model import Spreadsheet
from importer.run_bash_command import runrealcmd

class Preparation:

    @staticmethod
    def new_instance(spreadsheet: Spreadsheet, base: str, ticket: int, instance: int, excel_type: str):
        destination = "%s/%d" % (base, ticket)
        return Preparation(spreadsheet, destination, '%s/external_%d_%d.%s' % (destination, ticket, instance, excel_type))

    @staticmethod
    def new_instance_complete(spreadsheet: Spreadsheet, base: str, ticket: int, excel_type:str):
        destination = "%s/%d" % (base, ticket)
        return Preparation(spreadsheet, destination, '%s/complete_external_%d.%s' % (destination, ticket, excel_type))
        
    def __init__(self, spreadsheet: Spreadsheet, destination: str, spreadsheet_file: str):
        self.spreadsheet = spreadsheet
        self.destination = destination
        self.spreadsheet_file = spreadsheet_file

    def copy_files(self, source: str):
        for read in self.spreadsheet.reads:
            copyfile("%s/%s" % (source, read.forward_read), "%s/%s" % (self.destination, read.forward_read))
            if read.reverse_read is not None:
                copyfile("%s/%s" % (source, read.reverse_read), "%s/%s" % (self.destination, read.reverse_read))

    def download_files_from_ena(self,connections):
        reads_to_download=[read.forward_read for read in self.spreadsheet.reads if not self.check_if_file_downloaded(read.forward_read)]
        df = self.create_dataframe_from_list(reads_to_download)
        df = create_commands(df,connections, self.destination)
        submit_commands(df)

    def create_dataframe_from_list(self,reads):
        df = pd.DataFrame(([read, 'import_%s' % read] for read in reads),
                          columns=('Read accession', 'Job_name'))
        return df

    def check_if_file_downloaded(self, accession):
        single_ended_file = self.destination + '/' + accession + '.fastq.gz'
        forward_file= self.destination + '/' +  accession + '_1.fastq.gz'
        reverse_file= self.destination + '/' + accession + '_2.fastq.gz'
        if path.exists(single_ended_file):
            return True
        elif path.exists(forward_file) and path.exists(reverse_file):
                return True
        else:
            return False

    def save_workbook(self, workbook):
        workbook.save(self.spreadsheet_file)

    def create_destination_directory(self):
        if not path.isdir(self.destination):
            mkdir(self.destination)


def create_commands(df, connections, destination):
    for i in range(len(df)):
        df.loc[i,'enaDataGet_command'] = 'enaDataGet -f fastq -d %s %s' % (
            destination, df.loc[i, 'Read accession'])
        df.loc[i,'extract_data_command'] = 'mv %s/%s/* %s  && rm -rf %s/%s' % (
            destination, df.loc[i, 'Read accession'],destination,destination, df.loc[i, 'Read accession'])
        memory = "-M2000 -R 'select[mem>2000] rusage[mem=2000]' "
        if i >= connections:
            df.loc[i, 'Job_to_depend_on'] = df.loc[i - connections, 'Job_name']
            df.loc[i, 'Command'] = 'bsub -o %s/%s.o -e %s/%s.e %s -J import_%s -w \'ended("%s")\' "%s && %s"' % (
                destination, df.loc[i, 'Read accession'], destination, df.loc[i, 'Read accession'], memory, df.loc[i, 'Read accession'],df.loc[i, 'Job_to_depend_on'],df.loc[i,'enaDataGet_command'], df.loc[i,'extract_data_command'])
        if i< connections:
            df.loc[
                i, 'Command'] = 'bsub -o %s/%s.o -e %s/%s.e %s -J import_%s "%s && %s"' % (
            destination, df.loc[i, 'Read accession'], destination, df.loc[i, 'Read accession'],
            memory, df.loc[i, 'Read accession'], df.loc[i,'enaDataGet_command'], df.loc[i,'extract_data_command'])
    return df

def submit_commands(df):
    if not df.empty:
        df['download_return_code'] = df['Command'].apply(lambda x: runrealcmd(x))
    else:
        print ('No new files found to be downloaded')

class OutputSpreadsheetGeneratorXLS:

    FILE_ENDING = "xls"

    def __init__(self, spreadsheet: Spreadsheet, current_position: int):
        self.spreadsheet = spreadsheet
        self.row = current_position
        self.status_closed = False
        self.workbook = xlwt.Workbook()
        self.sheet = self.workbook.add_sheet('Sheet1')

    def build(self, breakpoint: int, download):
        self.build_import_info()
        self.build_read_headers()
        self.build_read_data(breakpoint if breakpoint > 0 else sys.maxsize, download)
        return self.workbook, self.status_closed, self.row

    def build_import_info(self):

        self.write_string(0, 'Supplier Name', self.spreadsheet.supplier)
        self.write_string(1, 'Supplier Organisation', self.spreadsheet.organisation)
        self.write_string(2, 'Sanger Contact Name', self.spreadsheet.contact)
        self.write_string(3, 'Sequencing Technology', self.spreadsheet.technology)
        self.write_string(4, 'Study Name', self.spreadsheet.name)
        self.write_string(5, 'Study Accession number', self.spreadsheet.accession)
        self.write_string(6, 'Total size of files in GBytes', self.spreadsheet.size)
        self.write_string(7, 'Data to be kept until', self.spreadsheet.limit)


    def build_read_data(self, breakpoint: int, download):
        for read in range(breakpoint):
            position = read + 10
            end_reached = self.row == len(self.spreadsheet.reads)
            if end_reached:
                self.status_closed = True
                break
            current_row = self.spreadsheet.reads[self.row]
            if download:
                if current_row.reverse_read == 'T':
                    forward_read_file = current_row.forward_read + '_1.fastq.gz'
                    self.sheet.write(position, 0, forward_read_file)
                    reverse_read_file = current_row.forward_read + '_2.fastq.gz'
                    self.sheet.write(position, 1, reverse_read_file)
                elif current_row.reverse_read == 'F':
                    forward_read_file = current_row.forward_read + '.fastq.gz'
                    self.sheet.write(position, 0, forward_read_file)
                    self.sheet.write(position, 1, '')
                else:
                    print('WARNING: some lines have invalid entries for the double-ended column (not T/F)')
            else:
                self.sheet.write(position, 0, current_row.forward_read)
                if current_row.reverse_read is not None:
                    self.sheet.write(position, 1, current_row.reverse_read)
            self.sheet.write(position, 2, current_row.sample_name)
            self.sheet.write(position, 4, current_row.taxon_id)
            self.sheet.write(position, 5, current_row.library_name)
            self.row += 1
        if self.row == len(self.spreadsheet.reads):
            self.status_closed = True

    def build_read_headers(self):
        index = 0
        for header in ['Filename', 'Mate File', 'Sample Name', 'Sample Accession number', 'Taxon ID', 'Library Name',
                       'Fragment Size', 'Read Count', 'Base Count', 'Comments']:
            self.sheet.write(9, index, header)
            index += 1

    def not_applicable(self, row, title):
        self.write_string(row, title, 'Unused field')

    def write_string(self, row, title, value):
        self.sheet.write(row, 0, title)
        self.sheet.write(row, 1, value)

class OutputSpreadsheetGeneratorXLSX:

    FILE_ENDING = "xlsx"

    def __init__(self, spreadsheet: Spreadsheet, current_position: int):
        self.spreadsheet = spreadsheet
        self.row = current_position
        self.status_closed = False
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.worksheets[0]

    def build(self, breakpoint: int, download):
        self.build_import_info()
        self.build_read_headers()
        self.build_read_data(breakpoint if breakpoint > 0 else sys.maxsize, download)
        return self.workbook, self.status_closed, self.row

    def build_import_info(self):

        self.write_string(1, 'Supplier Name', self.spreadsheet.supplier)
        self.write_string(2, 'Supplier Organisation', self.spreadsheet.organisation)
        self.write_string(3, 'Sanger Contact Name', self.spreadsheet.contact)
        self.write_string(4, 'Sequencing Technology', self.spreadsheet.technology)
        self.write_string(5, 'Study Name', self.spreadsheet.name)
        self.write_string(6, 'Study Accession number', self.spreadsheet.accession)
        self.write_string(7, 'Total size of files in GBytes', self.spreadsheet.size)
        self.write_string(8, 'Data to be kept until', self.spreadsheet.limit)


    def build_read_data(self, breakpoint: int, download):
        for read in range(breakpoint):
            position = read + 11
            end_reached = self.row == len(self.spreadsheet.reads)
            if end_reached:
                self.status_closed = True
                break
            current_row = self.spreadsheet.reads[self.row]
            if download:
                if current_row.reverse_read == 'T':
                    forward_read_file = current_row.forward_read + '_1.fastq.gz'
                    self.sheet.cell(row=position, column=1).value = forward_read_file
                    reverse_read_file = current_row.forward_read + '_2.fastq.gz'
                    self.sheet.cell(row=position, column=2).value = reverse_read_file
                elif current_row.reverse_read == 'F':
                    forward_read_file = current_row.forward_read + '.fastq.gz'
                    self.sheet.cell(row=position, column=1).value = forward_read_file
                    self.sheet.cell(row=position, column=2).value = ''
                else:
                    print('WARNING: some lines have invalid entries for the double-ended column (not T/F)')
            else:
                self.sheet.cell(row=position, column=1).value = current_row.forward_read
                if current_row.reverse_read is not None:
                    self.sheet.cell(row=position, column=2).value = current_row.reverse_read
            self.sheet.cell(row=position, column=3).value = current_row.sample_name
            self.sheet.cell(row=position, column=5).value = current_row.taxon_id
            self.sheet.cell(row=position, column=6).value = current_row.library_name
            self.row += 1
        if self.row == len(self.spreadsheet.reads):
            self.status_closed = True

    def build_read_headers(self):
        index = 0
        for header in ['Filename', 'Mate File', 'Sample Name', 'Sample Accession number', 'Taxon ID', 'Library Name',
                       'Fragment Size', 'Read Count', 'Base Count', 'Comments']:
            self.sheet.cell(row=10, column=(index+1)).value = header
            index += 1

    def not_applicable(self, row, title):
        self.write_string(row, title, 'Unused field')

    def write_string(self, row, title, value):
        self.sheet.cell(row=row, column=1).value = title
        self.sheet.cell(row=row, column=2).value = value
